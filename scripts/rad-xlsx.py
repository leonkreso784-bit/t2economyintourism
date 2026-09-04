# -*- coding: utf-8 -*-
"""
rad-xlsx.py — ANALIZA RADA u Excelu: koliko radimo, koliko dobro radimo, gdje smo s planovima.

Pokreni:  python scripts/rad-xlsx.py            (Windows: PYTHONUTF8=1 python scripts/rad-xlsx.py)
Traži:    openpyxl==3.1.5   →  pip install --user -r scripts/requirements-rad.txt
Izlaz:    docs/records/RAD.xlsx

ZAŠTO SKRIPTA, A NE SAMO DATOTEKA (Leon, 2026-09-05: „jednu excel tablicu koju ćemo stalno
nadopunjavat i imat ćemo nekoliko grafova"): openpyxl pri otvaranju postojeće knjige GUBI grafove —
svaka dopuna bez skripte bi ih obrisala. Zato su PODATKOVNI listovi izvor, a sažetak i grafovi se
svaki put grade iznova iz njih. Dopuna = pokreni skriptu (ona sama pročita git i PROGRESS.md).

KAD SE POKREĆE: na kraju svake faze (RASPORED §1: „zastanak na kraju faze"). U tablicu ulazi
SAMO ono što je odrađeno — commit je odrađen, PROGRESS-unos je odrađen; namjere i razgovori ne.

LISTOVI
  Dnevnik   — jedan red po commitu od datuma OD (git je jedini izvor: datum, vrijeme, opis,
              datoteke, +/− redaka, redaka u testovima i branama). „vrsta (auto)" je heuristika iz
              opisa commita; „vrsta (ručno)" je Leonov/naš ispravak i ČUVA SE pri osvježavanju.
  Isporuke  — jedan red po unosu u docs/records/PROGRESS.md (jedinica priče, ne koda).
  Faze      — zatvorene faze (konstante dolje, s datumima iz gita) + RASPORED F1–F7 (broj cigli i
              ✅ se ČITAJU iz docs/plan/RASPORED.md, ne prepisuju).
  Vizije    — planovi i vizije sa stanjem i procjenom postotka. RUČNI list: čuva se kakav jest;
              sjeme se upiše samo kad lista nema.
  Sažetak   — tempo po danu · vrste rada · kvaliteta · faze — i SVI grafovi. Uvijek se pregrađuje.
  Upute     — kako se knjiga puni, što znače stupci, što je proxy a što mjera.

ŠTO JE MJERA, A ŠTO PROXY: commiti, redci, datumi, broj cigli = mjera (iz gita/dokumenata).
„Sati" = raspon od prvog do zadnjeg commita u danu — proxy za trajanje rada, ne štoperica.
Postotak vizije = procjena, i tako je označen.
"""
import datetime as dt
import io
import os
import re
import subprocess
import sys
from collections import OrderedDict, defaultdict

try:
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit('Treba openpyxl 3.1.5:  pip install --user -r scripts/requirements-rad.txt')

KORIJEN = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
IZLAZ = os.path.join(KORIJEN, 'docs', 'records', 'RAD.xlsx')
PROGRESS = os.path.join(KORIJEN, 'docs', 'records', 'PROGRESS.md')
RASPORED = os.path.join(KORIJEN, 'docs', 'plan', 'RASPORED.md')

# Od kada Dnevnik broji (Leon, 2026-09-05: „napravi u zadnjih tjedan dana"). Ranije = HISTORY.md.
OD = '2026-08-29'

VRSTE = ['planiranje', 'vođenje dokumentacije', 'izvođenje procesa', 'poliranje koda', 'debugging']

# Zatvorene faze — konstante (datumi potvrđeni u gitu: prvi/zadnji commit s oznakom faze).
# ⚠️ Broj cigli zatvorene faze se NE upisuje rukom nego BROJI: cigla = commit čiji opis počinje
# oznakom faze (regex dolje). Prva verzija je nosila procjene („39/35") — procjena nije mjera.
# Aktivne faze (RASPORED F1–F7) se čitaju iz RASPORED.md i NE pišu ovdje.
ZATVORENE_FAZE = [
    # ime, početak, kraj, regex oznake u opisu commita, napomena
    ('Osobni UGC-graditelj F1–F5', '2026-08-02', '2026-08-06', r'^(F[1-5][: ]|Merge F5|docs\(F5\))',
     'DB temelj → moji materijali → editor u čvoru → privatne slike → produkcija (HISTORY.md)'),
    ('Frontend redizajn C0–C7 + KOSTUR·TELEFON·POLICA·SEO', '2026-08-07', '2026-09-01',
     r'^(C\d[ab]?|T\d|K\d|N\d|S\d|ALAT-\d)', 'Tailwind CLI, tokeni, 3 teme, birač; C3/C5b/C6/C7 dug → RASPORED F4 (BACKLOG)'),
    ('MREŽA — sanacija A–E', '2026-08-31', '2026-09-01', r'^(faza MREZA|MREZA)',
     'CSP enforce, RLS po retku, check:i18n/tokens/final/cascade, CI shardanje, D1–D4, E1–E4'),
    ('RAČUN R1 — OAuth + upitnik + dijalog', '2026-09-02', '2026-09-02', r'^R1',
     'Google uživo potvrđen, U1/U2/U5 popravci, FB maknut (čeka Metine ključeve); R2/R3 → F2'),
]

# „Sati rada" — git-hours heuristika: dva uzastopna commita razmaknuta manje od PRAG sati pripadaju
# istoj sesiji i razmak se broji kao rad; prvi commit sesije dobiva POCETAK sati (rad prije njega).
# Raspon prvog i zadnjeg commita u danu je bio prva verzija — noć između dva commita brojio je kao rad.
PRAG_SESIJE_H = 2.0
POCETAK_SESIJE_H = 0.5

# Sjeme lista „Vizije" — upisuje se SAMO ako list ne postoji (list je ručni, čuva se).
VIZIJE_SJEME = [
    # stavka, izvor, stanje, % (procjena), napomena
    ('UGC platforma za SVE — osobni graditelj (privatno)', 'ADR-024/029, CREATE_BACKEND_SPEC', 'isporučeno', 100,
     'Na produkciji 2026-08-06; objava/dijeljenje je zaseban plan (F7)'),
    ('Frontend redizajn — Apple smjer, tokeni, teme', 'FRONTEND_REDIZAJN (arhiv)', 'isporučeno', 85,
     'Na produkciji 2026-09-01; ostatak = CSS dug C3/C5b/C6/C7 (6088 redaka) → F4'),
    ('MREŽA — sanacija sigurnosti i brana', 'MREZA (arhiv)', 'isporučeno', 100, 'Na produkciji 2026-09-01'),
    ('Računi — email+lozinka, cloud-sync, profil', 'PRD', 'isporučeno', 100, 'LIVE od lipnja; GDPR brisanje kroz Edge Function'),
    ('Računi — Google prijava (R1)', 'RACUN (arhiv)', 'isporučeno', 100, 'Na produkciji 2026-09-02; Facebook čeka Metine ključeve'),
    ('Računi — tema uz račun, slika, uređivanje, mail-obavijesti (F2)', 'RASPORED F2', 'planirano', 0, 'Iza F1'),
    ('Stranica = izgled maila, crno na tamnom uređaju (F1)', 'RASPORED F1', 'pokrenuto', 33,
     'F1/1–F1/3 na grani; ostaje živa brana, 4 stranice bez teme, trzanje, hover, Tinder-špil'),
    ('Dvojezičnost sučelja HR/EN — 421 → 0 (F3)', 'ADR-033, RASPORED F3', 'planirano', 15,
     'Mehanizam i brana check:i18n isporučeni (MREŽA B5); prijevod čeka'),
    ('Čišćenje CSS-duga do nule (F4)', 'RASPORED F4', 'planirano', 0, 'css:debt 6088 redaka, 2 !important'),
    ('Vježbe kao PODATAK — knjižnica recepata (F5)', 'RASPORED F5, BUG-012', 'planirano', 0,
     'Izmjereno: 65 % vježbi je već čisti podatak; prvo prebrojiti recepte'),
    ('MCP cjevovod — AI korisnika stvara gradivo kroz chat (F6)', 'ADR-030/031, RASPORED F6', 'pokrenuto', 10,
     'Read-only spike (mcp-admin, untracked); konektor traži OAuth (F2)'),
    ('Objava i dijeljenje materijala + povijest učenja (F7)', 'RASPORED F7', 'planirano', 0, 'Daje smisao povijesti i grafikonima'),
    ('HR-ekspanzija — 3 smjera FMTU, HR u Supabase', 'ADR-022, docs/subjects', 'pokrenuto', 50,
     '7 HR predmeta live; 4 kvantitativna BEZ VLASNIKA (suradnik otkazan 2026-09-04)'),
    ('AI tutor — plaćeni ili „donesi svoj ključ"', 'VISION F1/F6', 'ideja', 0, 'Iza F7; trošak neizmjeren'),
    ('Natjecanje i društveno — ljestvice, statistika', 'ROADMAP M3', 'ideja', 0, ''),
    ('Monetizacija — funkcionalnost, ne sadržaj', 'ROADMAP M4, MONETIZATION', 'ideja', 0, 'Tek na skali'),
    ('Spaced repetition', 'RASPORED §5', 'parkirano', 0, 'Nije otkazan, samo nije sljedeći'),
    ('Matura (NCVVO materijali)', 'RASPORED §5', 'parkirano', 0, 'Leon: „neću otvarat maturu" — pravno pitanje'),
    ('Simulacija vođenja hotela', 'ideas/HOTEL_SIM', 'ideja', 0, 'Zaseban proizvod, posuđuje primitive'),
    ('Seoba sa Supabasea (self-host)', 'BACKLOG §SELF-HOST', 'odbijeno', 0, 'Leon 2026-09-01: „Nastavlja Supabase do daljnjeg"'),
    ('Next.js / framework', 'ADR-028', 'odbijeno', 0, 'Vanilla JS bez build-koraka ostaje'),
]


# ───────────────────────── git ─────────────────────────
def git(*args):
    return subprocess.run(['git'] + list(args), cwd=KORIJEN, capture_output=True,
                          encoding='utf-8', errors='replace').stdout


OZNAKA = re.compile(r'^(F\d/\d|C\d[ab]?(?:/\d\w*)?|MREZA[- ]?[A-E]\d?(?: \(\d/\d\))?|R\d(?:/[A-Z0-9+]+)?|T\d|ALAT-\d|BUG-\d+|U\d)')


def vrsta_iz_opisa(s):
    """Heuristika po opisu commita. Redoslijed je bitan: planiranje prije dokumentacije
    (odluke i specovi su „docs:" commiti), debugging prije poliranja (fix na CSS-u je debugging)."""
    t = s.lower()
    if re.search(r'raspored|dobiva svoj spec|tracnice|tračnice|mjera za c|adr-0|odluk|presud|revizija je dala|sto sada|što sada|zadatka za sljedecu|zadatak za sljedecu', t):
        return 'planiranje'
    if re.search(r'^docs|compact|revizija pred|^r\d/docs|ishod --|deploy-zapis|zapis uz', t):
        return 'vođenje dokumentacije'
    if re.search(r'^fix|bug-\d|^ci:|popravak ci|kvar|obara|^test:|lagati|laze|laže|rupe|bomba|vise ne postoji|više ne postoji', t):
        return 'debugging'
    if re.search(r'^c[4-7]|paleta|mrtv|selektorski|\bvan\b|audit|node 24|migrira|utility|ljestv|raspusten|raspušten|skuplj|refactor|^alat|ugasen|ugašen|dug placen|dug plaćen', t):
        return 'poliranje koda'
    return 'izvođenje procesa'


def podvrsta(s):
    t = s.lower()
    if '🚀' in s or 'na produkciji' in t or 'deploy' in t:
        return 'deploy'
    if re.search(r'check:|brana|gate|^alat|^test|sonda|probe|mjera', t):
        return 'brane i mjerenje'
    if OZNAKA.match(s):
        return 'cigla'
    return 'ostalo'


def commiti_od(od):
    """Jedan zapis po commitu: datum, vrijeme, sha, opis, datoteke, +, −, testni redci."""
    izlaz = git('log', '--since=' + od, '--date=format:%Y-%m-%d %H:%M', '--reverse',
                '--format=@@%h|%ad|%s', '--numstat')
    zapisi, cur = [], None
    for line in izlaz.splitlines():
        if line.startswith('@@'):
            h, d, s = line[2:].split('|', 2)
            cur = {'sha': h, 'datum': d[:10], 'vrijeme': d[11:], 'opis': s, 'datoteka': 0,
                   'plus': 0, 'minus': 0, 'test': 0}
            zapisi.append(cur)
        elif cur and line.strip():
            dijelovi = line.split('\t')
            if len(dijelovi) == 3:
                p, m, put = dijelovi
                p = 0 if p == '-' else int(p)
                m = 0 if m == '-' else int(m)
                cur['datoteka'] += 1
                cur['plus'] += p
                cur['minus'] += m
                if put.startswith('tests/') or '/check-' in put or put.endswith('.test.js') or put.endswith('.spec.js'):
                    cur['test'] += p + m
    return zapisi


def commiti_u_razdoblju(od, do, uzorak):
    izlaz = git('log', '--since=' + od, '--until=' + do + ' 23:59', '--format=%s')
    r = re.compile(uzorak)
    return sum(1 for s in izlaz.splitlines() if r.search(s))


def sati_po_danu(commiti):
    """git-hours: razmak < PRAG_SESIJE_H između uzastopnih commita = rad; novi niz = + POCETAK_SESIJE_H.
    Sat se pripisuje danu commita koji ga zatvara."""
    sati = defaultdict(float)
    prethodni = None
    for z in commiti:
        t = dt.datetime.fromisoformat(z['datum'] + ' ' + z['vrijeme'])
        if prethodni is not None and (t - prethodni).total_seconds() / 3600 < PRAG_SESIJE_H:
            sati[z['datum']] += (t - prethodni).total_seconds() / 3600
        else:
            sati[z['datum']] += POCETAK_SESIJE_H
        prethodni = t
    return {d: round(h, 1) for d, h in sati.items()}


# ───────────────────────── dokumenti ─────────────────────────
def isporuke_iz_progressa(od):
    """Naslovi `## YYYY-MM-DD (MODEL) — naslov` iz PROGRESS.md od datuma OD."""
    if not os.path.exists(PROGRESS):
        return []
    tekst = io.open(PROGRESS, encoding='utf-8').read()
    zapisi = []
    for m in re.finditer(r'^## (\d{4}-\d{2}-\d{2})(?:\s*\(([^)]*)\))?\s*[—-]+\s*(.+)$', tekst, re.M):
        datum, model, naslov = m.group(1), (m.group(2) or '').strip(), m.group(3).strip()
        if datum < od:
            continue
        zapisi.append({'datum': datum, 'model': model.upper() if model else '', 'naslov': naslov,
                       'vrsta': vrsta_iz_opisa(naslov),
                       'deploy': bool('🚀' in naslov or re.search(r'deploy|main na `', naslov.lower()))})
    zapisi.sort(key=lambda z: z['datum'])
    return zapisi


def faze_iz_rasporeda():
    """Broj cigli i ✅ po fazi F1–F7 — ČITA se iz RASPORED.md."""
    faze = OrderedDict()
    if not os.path.exists(RASPORED):
        return faze
    tekst = io.open(RASPORED, encoding='utf-8').read()
    imena = dict(re.findall(r'^### (F\d) · ([^\n]+)$', tekst, re.M))
    for m in re.finditer(r'^\| \*\*(F\d)/(\d+)\*\*\s*(✅?)', tekst, re.M):
        f = m.group(1)
        z = faze.setdefault(f, {'ime': f + ' · ' + imena.get(f, '').strip(), 'ukupno': 0, 'gotovo': 0})
        z['ukupno'] += 1
        if m.group(3):
            z['gotovo'] += 1
    return faze


# ───────────────────────── stil ─────────────────────────
NASLOV = Font(bold=True, color='FFFFFF')
ISPUNA = PatternFill('solid', fgColor='2F3A4F')
MEKA = PatternFill('solid', fgColor='EEF2F8')
TANKA = Side(style='thin', color='D6DEE8')
OKVIR = Border(top=TANKA, bottom=TANKA, left=TANKA, right=TANKA)


def zaglavlje(ws, red, stupci, sirine=None):
    for i, ime in enumerate(stupci, 1):
        c = ws.cell(row=red, column=i, value=ime)
        c.font, c.fill, c.alignment, c.border = NASLOV, ISPUNA, Alignment(vertical='center', wrap_text=True), OKVIR
    if sirine:
        for i, w in enumerate(sirine, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[red].height = 30


def tablica(ws, red, stupci, redci, sirine=None, filtar=True):
    zaglavlje(ws, red, stupci, sirine)
    for r, vals in enumerate(redci, red + 1):
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = OKVIR
            cell.alignment = Alignment(vertical='top', wrap_text=isinstance(v, str) and len(v) > 40)
    if filtar and redci:
        ws.auto_filter.ref = '%s%d:%s%d' % ('A', red, get_column_letter(len(stupci)), red + len(redci))
    return red + len(redci) + 1


# ───────────────────────── glavno ─────────────────────────
def main():
    rucne_vrste, vizije_postojece = {}, None
    if os.path.exists(IZLAZ):
        stara = openpyxl.load_workbook(IZLAZ)
        if 'Dnevnik' in stara.sheetnames:
            ws = stara['Dnevnik']
            glave = [c.value for c in ws[1]]
            if 'sha' in glave and 'vrsta (ručno)' in glave:
                i_sha, i_r = glave.index('sha'), glave.index('vrsta (ručno)')
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[i_sha] and row[i_r]:
                        rucne_vrste[row[i_sha]] = row[i_r]
        if 'Vizije' in stara.sheetnames:
            ws = stara['Vizije']
            vizije_postojece = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]

    commiti = commiti_od(OD)
    isporuke = isporuke_iz_progressa(OD)
    faze_r = faze_iz_rasporeda()
    danas = dt.date.today().isoformat()

    wb = openpyxl.Workbook()

    # ── Sažetak (prvi list; puni se poslije podataka, ali kreiramo ga prvog radi redoslijeda)
    ws_s = wb.active
    ws_s.title = 'Sažetak'

    # ── Dnevnik
    ws = wb.create_sheet('Dnevnik')
    stupci = ['datum', 'vrijeme', 'sha', 'opis', 'oznaka', 'vrsta (auto)', 'vrsta (ručno)', 'podvrsta',
              'datoteka', '+ redaka', '− redaka', 'redaka u testovima/branama']
    redci = []
    for z in commiti:
        m = OZNAKA.match(z['opis'])
        redci.append([z['datum'], z['vrijeme'], z['sha'], z['opis'], m.group(1) if m else '',
                      vrsta_iz_opisa(z['opis']), rucne_vrste.get(z['sha'], ''), podvrsta(z['opis']),
                      z['datoteka'], z['plus'], z['minus'], z['test']])
    tablica(ws, 1, stupci, redci, [11, 7, 9, 70, 12, 20, 18, 16, 9, 9, 9, 12])
    ws.freeze_panes = 'E2'

    # ── Isporuke
    ws = wb.create_sheet('Isporuke')
    redci = [[i['datum'], i['model'], i['naslov'], i['vrsta'], 'da' if i['deploy'] else ''] for i in isporuke]
    tablica(ws, 1, ['datum', 'model', 'što je isporučeno (PROGRESS.md)', 'vrsta (auto)', 'deploy'], redci,
            [11, 8, 90, 20, 8])
    ws.freeze_panes = 'A2'

    # ── Faze
    ws = wb.create_sheet('Faze')
    redci = []
    for ime, od, do, uzorak, nap in ZATVORENE_FAZE:
        dani = (dt.date.fromisoformat(do) - dt.date.fromisoformat(od)).days + 1
        n = commiti_u_razdoblju(od, do, uzorak)   # cigla zatvorene faze = commit s oznakom faze
        redci.append([ime, 'zatvoreno', od, do, dani, n, n, 1.0, round(n / dani, 2), n, round(n / dani, 1),
                      nap + ' · cigla = commit s oznakom faze'])
    f1_pocetak = None
    for z in commiti:
        if z['opis'].startswith('F1/'):
            f1_pocetak = z['datum']
            break
    for f, z in faze_r.items():
        stanje = 'u tijeku' if 0 < z['gotovo'] < z['ukupno'] else ('zatvoreno' if z['ukupno'] and z['gotovo'] == z['ukupno'] else 'planirano')
        od = f1_pocetak if f == 'F1' else ''
        dani = (dt.date.today() - dt.date.fromisoformat(od)).days + 1 if od else ''
        n = commiti_u_razdoblju(od, danas, r'^' + f + '/') if od else 0
        redci.append([z['ime'], stanje, od, '' if stanje != 'zatvoreno' else danas, dani, z['ukupno'], z['gotovo'],
                      round(z['gotovo'] / z['ukupno'], 2) if z['ukupno'] else 0,
                      round(z['gotovo'] / dani, 2) if dani else '', n, round(n / dani, 1) if dani else '',
                      'čita se iz RASPORED.md'])
    tablica(ws, 1, ['faza', 'stanje', 'početak', 'kraj', 'dana', 'cigle', 'gotove', 'udio', 'cigle/dan',
                    'commiti', 'commiti/dan', 'napomena'], redci, [46, 10, 11, 11, 6, 7, 7, 7, 9, 8, 10, 60], filtar=False)
    ws.freeze_panes = 'B2'
    faze_redci = redci

    # ── Vizije (ručni list)
    ws = wb.create_sheet('Vizije')
    redci = [list(r[:5]) for r in vizije_postojece] if vizije_postojece else [list(v) for v in VIZIJE_SJEME]
    tablica(ws, 1, ['plan / vizija', 'izvor', 'stanje', '% (procjena)', 'napomena'], redci, [56, 30, 12, 12, 70])
    ws.freeze_panes = 'A2'
    vizije_redci = redci

    # ── Sažetak: tablice + grafovi
    ws = ws_s
    ws.column_dimensions['A'].width = 34
    for col in 'BCDEFGH':
        ws.column_dimensions[col].width = 13
    ws['A1'] = 'Sokrat Study — analiza rada'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'Generirano %s iz gita (od %s), PROGRESS.md i RASPORED.md. Pokreni: python scripts/rad-xlsx.py' % (danas, OD)
    ws['A2'].font = Font(italic=True, color='666666')

    # tempo po danu
    po_danu = OrderedDict()
    sati_dan = sati_po_danu(commiti)
    for z in commiti:
        d = po_danu.setdefault(z['datum'], {'n': 0, 'redci': 0, 'test': 0})
        d['n'] += 1
        d['redci'] += z['plus'] + z['minus']
        d['test'] += z['test']
    isp_po_danu = defaultdict(int)
    dep_po_danu = defaultdict(int)
    for i in isporuke:
        isp_po_danu[i['datum']] += 1
        if i['deploy']:
            dep_po_danu[i['datum']] += 1
    red = 4
    ws.cell(row=red, column=1, value='TEMPO PO DANU').font = Font(bold=True)
    red += 1
    tempo_redci = []
    kum = 0
    for datum, d in po_danu.items():
        kum += d['n']
        tempo_redci.append([datum, d['n'], kum, d['redci'], sati_dan.get(datum, 0), isp_po_danu.get(datum, 0), dep_po_danu.get(datum, 0), d['test']])
    tempo_start = red
    red = tablica(ws, red, ['dan', 'commiti', 'commiti kumulativno', 'redaka (+/−)', 'sati (git-hours proxy)',
                            'isporuke (PROGRESS)', 'deploya', 'redaka u testovima'], tempo_redci, filtar=False)
    tempo_kraj = red - 1

    # vrste rada
    red += 1
    ws.cell(row=red, column=1, value='VRSTE RADA (po commitima; „ručno" pregazi „auto")').font = Font(bold=True)
    red += 1
    po_vrsti = OrderedDict((v, {'n': 0, 'redci': 0}) for v in VRSTE)
    for z in commiti:
        v = rucne_vrste.get(z['sha']) or vrsta_iz_opisa(z['opis'])
        if v not in po_vrsti:
            po_vrsti[v] = {'n': 0, 'redci': 0}
        po_vrsti[v]['n'] += 1
        po_vrsti[v]['redci'] += z['plus'] + z['minus']
    ukupno_n = sum(d['n'] for d in po_vrsti.values()) or 1
    vrste_redci = [[v, d['n'], round(d['n'] / ukupno_n, 3), d['redci']] for v, d in po_vrsti.items()]
    vrste_start = red
    red = tablica(ws, red, ['vrsta rada', 'commita', 'udio', 'redaka (+/−)'], vrste_redci, filtar=False)
    for r in range(vrste_start + 1, red):
        ws.cell(row=r, column=3).number_format = '0.0%'
    vrste_kraj = red - 1

    # kvaliteta i brzina
    red += 1
    ws.cell(row=red, column=1, value='KVALITETA I BRZINA (razdoblje od %s)' % OD).font = Font(bold=True)
    red += 1
    dana = len(po_danu) or 1
    n_comm = len(commiti)
    n_fix = sum(1 for z in commiti if (rucne_vrste.get(z['sha']) or vrsta_iz_opisa(z['opis'])) == 'debugging')
    n_docs = sum(1 for z in commiti if (rucne_vrste.get(z['sha']) or vrsta_iz_opisa(z['opis'])) == 'vođenje dokumentacije')
    n_test = sum(z['test'] for z in commiti)
    n_dep = sum(dep_po_danu.values())
    n_ci = sum(1 for z in commiti if re.search(r'^ci:|popravak ci|job je otkazan', z['opis'].lower()))
    n_leon = sum(1 for i in isporuke if 'leon' in i['naslov'].lower())
    sati = sum(r[4] for r in tempo_redci)
    pokazatelji = [
        ['radnih dana (dana s commitom)', dana, 'mjera'],
        ['commita', n_comm, 'mjera'],
        ['commita po radnom danu', round(n_comm / dana, 1), 'mjera'],
        ['isporuka (PROGRESS unosa)', len(isporuke), 'mjera'],
        ['isporuka po radnom danu', round(len(isporuke) / dana, 1), 'mjera'],
        ['sati rada (git-hours proxy)', round(sati, 1), 'PROXY — razmaci < %g h među commitima + %g h po sesiji, ne štoperica' % (PRAG_SESIJE_H, POCETAK_SESIJE_H)],
        ['commita po satu (proxy)', round(n_comm / sati, 1) if sati else '', 'proxy'],
        ['redaka promijenjeno (+/−)', sum(z['plus'] + z['minus'] for z in commiti), 'mjera'],
        ['redaka u testovima i branama', n_test, 'mjera — koliko rada ide u dokaz, ne u tvrdnju'],
        ['udio testnih redaka', round(n_test / max(1, sum(z['plus'] + z['minus'] for z in commiti)), 3), 'mjera'],
        ['deploya na produkciju', n_dep, 'mjera (🚀 u PROGRESS.md)'],
        ['debugging commita', n_fix, 'mjera — koliko se vraćamo na već napravljeno'],
        ['udio debugginga', round(n_fix / max(1, n_comm), 3), 'niže = bolje, ali 0 znači da brane ne rade'],
        ['udio dokumentacije', round(n_docs / max(1, n_comm), 3), 'režija: pravilo #3 + #6 (audit prije compacta)'],
        ['CI-padova popravljenih', n_ci, 'mjera'],
        ['isporuka pokrenutih Leonovim nalazom', n_leon, 'mjera — koliko smjera dolazi s uređaja, ne iz plana'],
        ['zatvorenih faza u razdoblju', sum(1 for r in faze_redci if r[1] == 'zatvoreno' and r[3] >= OD), 'mjera'],
        ['prosječno trajanje zatvorene faze (dana)', round(sum(r[4] for r in faze_redci if r[1] == 'zatvoreno') / max(1, sum(1 for r in faze_redci if r[1] == 'zatvoreno')), 1), 'mjera — sve zatvorene faze na listu Faze'],
    ]
    kval_start = red
    red = tablica(ws, red, ['pokazatelj', 'vrijednost', 'što je to'], pokazatelji, filtar=False)
    for r in range(kval_start + 1, red):
        if 'udio' in str(ws.cell(row=r, column=1).value):
            ws.cell(row=r, column=2).number_format = '0.0%'

    # faze — sažeto za graf
    red += 1
    ws.cell(row=red, column=1, value='FAZE — gotovo / preostalo (cigle)').font = Font(bold=True)
    red += 1
    faze_kratko = [[r[0][:40], r[6], max(0, r[5] - r[6]), r[1]] for r in faze_redci]
    faze_start = red
    red = tablica(ws, red, ['faza', 'gotove cigle', 'preostale cigle', 'stanje'], faze_kratko, filtar=False)
    faze_kraj = red - 1

    # vizije — sažeto
    red += 1
    ws.cell(row=red, column=1, value='PLANOVI I VIZIJE — po stanju').font = Font(bold=True)
    red += 1
    po_stanju = OrderedDict()
    for v in vizije_redci:
        po_stanju.setdefault(v[2], 0)
        po_stanju[v[2]] += 1
    viz_start = red
    red = tablica(ws, red, ['stanje', 'stavki'], [[k, n] for k, n in po_stanju.items()], filtar=False)
    viz_kraj = red - 1

    # ── grafovi (svi u stupcu J nadalje)
    def bar(naslov, ref_data, ref_cats, sidro, stacked=False, w=18, h=8, tip='col'):
        ch = BarChart()
        ch.type = tip
        ch.title = naslov
        ch.add_data(ref_data, titles_from_data=True)
        ch.set_categories(ref_cats)
        if stacked:
            ch.grouping, ch.overlap = 'stacked', 100
        ch.width, ch.height = w, h
        ws.add_chart(ch, sidro)
        return ch

    kat = Reference(ws, min_col=1, min_row=tempo_start + 1, max_row=tempo_kraj)
    bar('Koliko radimo — commita po danu', Reference(ws, min_col=2, min_row=tempo_start, max_row=tempo_kraj), kat, 'J4')
    ln = LineChart()
    ln.title = 'Tempo — commiti kumulativno'
    ln.add_data(Reference(ws, min_col=3, min_row=tempo_start, max_row=tempo_kraj), titles_from_data=True)
    ln.set_categories(kat)
    ln.width, ln.height = 18, 8
    ws.add_chart(ln, 'T4')
    bar('Redaka promijenjeno po danu (+/−)', Reference(ws, min_col=4, min_row=tempo_start, max_row=tempo_kraj), kat, 'J21')
    bar('Sati rada po danu (git-hours proxy)', Reference(ws, min_col=5, min_row=tempo_start, max_row=tempo_kraj), kat, 'T21')

    pie = PieChart()
    pie.title = 'Vrste rada — udio commita'
    pie.add_data(Reference(ws, min_col=2, min_row=vrste_start, max_row=vrste_kraj), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=vrste_start + 1, max_row=vrste_kraj))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.width, pie.height = 18, 9
    ws.add_chart(pie, 'J38')
    bar('Vrste rada — redaka (+/−)', Reference(ws, min_col=4, min_row=vrste_start, max_row=vrste_kraj),
        Reference(ws, min_col=1, min_row=vrste_start + 1, max_row=vrste_kraj), 'T38', tip='bar')

    bar('Faze — gotove i preostale cigle', Reference(ws, min_col=2, max_col=3, min_row=faze_start, max_row=faze_kraj),
        Reference(ws, min_col=1, min_row=faze_start + 1, max_row=faze_kraj), 'J57', stacked=True, w=26, h=10, tip='bar')
    bar('Planovi i vizije — po stanju', Reference(ws, min_col=2, min_row=viz_start, max_row=viz_kraj),
        Reference(ws, min_col=1, min_row=viz_start + 1, max_row=viz_kraj), 'T78', w=18, h=8)

    # ── Upute
    ws = wb.create_sheet('Upute')
    ws.column_dimensions['A'].width = 120
    upute = [
        ('KAKO SE KNJIGA PUNI', True),
        ('1. Na kraju svake faze (RASPORED §1) pokreni:  python scripts/rad-xlsx.py   (Windows: PYTHONUTF8=1 ispred).', False),
        ('2. Skripta sama pročita git od %s, PROGRESS.md i RASPORED.md, pa PREGRADI Sažetak i sve grafove.' % OD, False),
        ('3. U tablicu ulazi SAMO ono što je odrađeno: commit i PROGRESS-unos jesu; namjera, razgovor i plan nisu.', False),
        ('', False),
        ('ŠTO SE ČUVA PRI OSVJEŽAVANJU', True),
        ('• Dnevnik → stupac „vrsta (ručno)": upiši ručno kad heuristika krivo presudi; čuva se po SHA.', False),
        ('• Vizije → cijeli list je ručni (stavka, izvor, stanje, %, napomena). Sjeme se upiše samo kad lista nema.', False),
        ('• Sve ostalo se računa iznova. Ne uređuj Sažetak, Dnevnik (osim „ručno") ni Faze — sljedeće osvježavanje ih pregazi.', False),
        ('', False),
        ('ŠTO ZNAČE STUPCI', True),
        ('• vrsta rada: planiranje · vođenje dokumentacije · izvođenje procesa · poliranje koda · debugging (Leonovih pet).', False),
        ('  „auto" je heuristika po opisu commita (redoslijed: planiranje > dokumentacija > debugging > poliranje > izvođenje).', False),
        ('• podvrsta: cigla (commit s oznakom faze) · brane i mjerenje · deploy · ostalo.', False),
        ('• redaka u testovima/branama: +/− redaka u tests/**, scripts/check-*, *.test.js, *.spec.js — koliko rada ide u DOKAZ.', False),
        ('• sati (git-hours proxy): razmak među uzastopnim commitima kraći od %g h broji se kao rad, svaka nova sesija dobiva %g h. PROXY, ne štoperica.' % (PRAG_SESIJE_H, POCETAK_SESIJE_H), False),
        ('• Faze: zatvorene su konstante u skripti (datumi iz gita), a njihova cigla = commit s oznakom faze (mjera, ne procjena); F1–F7 se ČITAJU iz RASPORED.md (✅ = gotova cigla).', False),
        ('• Vizije %: PROCJENA, tako i piše. Stanja: isporučeno · pokrenuto · planirano · ideja · parkirano · odbijeno.', False),
        ('', False),
        ('KAKO ČITATI „KOLIKO DOBRO RADIMO"', True),
        ('• udio testnih redaka gore = više dokaza po redu koda; udio debugginga niže = manje vraćanja — ali 0 znači da brane ne love.', False),
        ('• udio dokumentacije = režija pravila #3 i #6; visok je namjerno (ADR-027: znanje u kod i zapise), prati trend, ne broj.', False),
        ('• isporuke pokrenute Leonovim nalazom: koliko smjera dolazi s uređaja, a ne iz plana — visok broj = plan zaostaje za stvarnošću.', False),
        ('• cigle/dan i commiti/dan po fazi = brzina; uspoređuj faze iste vrste (CSS-faza nije račun-faza).', False),
    ]
    for i, (t, b) in enumerate(upute, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(bold=b)
        c.alignment = Alignment(wrap_text=True, vertical='top')

    os.makedirs(os.path.dirname(IZLAZ), exist_ok=True)
    wb.save(IZLAZ)
    print('✅ %s  —  %d commita · %d isporuka · %d faza · %d vizija · grafova: 8' % (
        os.path.relpath(IZLAZ, KORIJEN), len(commiti), len(isporuke), len(faze_redci), len(vizije_redci)))


if __name__ == '__main__':
    main()
